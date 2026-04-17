
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os

# ─── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="A&DS Monthly Meeting",
    page_icon="🗓️",
    layout="centered"
)

# ─── Clorox Brand Colors & Custom CSS ───────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    .stApp {
        background-color: #F4F6FB;
    }

    /* Hero Header Banner */
    .hero-banner {
        background: linear-gradient(135deg, #5DB8FD 0%, #5DB8FD 100%);
        padding: 40px 36px 32px 36px;
        border-radius: 18px;
        margin-bottom: 28px;
        box-shadow: 0 8px 32px rgba(93, 184, 253, 0.25);
        text-align: center;
        position: relative;
        overflow: hidden;
    }
    .hero-banner::before {
        content: "";
        position: absolute;
        top: -30px; right: -30px;
        width: 160px; height: 160px;
        background: rgba(93, 184, 253, 0.12);
        border-radius: 50%;
    }
    .hero-banner::after {
        content: "";
        position: absolute;
        bottom: -40px; left: -20px;
        width: 200px; height: 200px;
        background: rgba(93, 184, 253, 0.07);
        border-radius: 50%;
    }

    .hero-tag {
        display: inline-block;
        background-color: #FFFFFF;   /* White background */
        color: #5DB8FD;              /* Blue text */
        border: 2px solid #5DB8FD;   /* Blue border */
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        padding: 4px 14px;
        border-radius: 20px;
        margin-bottom: 14px;
    }

    .hero-title {
        color: #FFFFFF;
        font-size: 2.2rem;
        font-weight: 800;
        margin: 0 0 6px 0;
        line-height: 1.2;
        letter-spacing: -0.5px;
    }
    .hero-subtitle {
        color: rgba(255,255,255,0.80);
        font-size: 1.0rem;
        font-weight: 400;
        margin: 0;
    }
    .hero-accent {
        color: #FFD100;
        font-weight: 700;
    }

    /* Form Card */
    .form-card {
        background: #FFFFFF;
        border-radius: 16px;
        padding: 32px 36px;
        box-shadow: 0 2px 20px rgba(0,0,0,0.07);
        margin-bottom: 20px;
    }
    .section-label {
        font-size: 0.98rem;
        font-weight: 700;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        color: #5DB8FD;
        margin-bottom: 8px;
        display: block;
    }
    .divider {
        border: none;
        border-top: 2px solid #F0F2F8;
        margin: 22px 0;
    }
    .diet-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 10px;
        margin-top: 8px;
    }
    .diet-chip {
        background: #F0F4FF;
        border: 1.5px solid #C5D3F5;
        border-radius: 10px;
        padding: 10px 14px;
        font-size: 0.92rem;
        color: #5DB8FD;
        cursor: pointer;
        transition: all 0.2s;
    }
    .diet-chip:hover {
        background: #5DB8FD;
        color: white;
    }

    /* Submit Button */
    .stButton > button {
        background: #5DB8FD !important;   /* Blue background */
        color: #000000 !important;        /* Black text */
        font-weight: 700 !important;
        font-size: 1.05rem !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 14px 0 !important;
        width: 100% !important;
        letter-spacing: 0.5px !important;
        box-shadow: 0 4px 16px rgba(93, 184, 253, 0.30) !important;
        transition: all 0.2s !important;
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(93, 184, 253, 0.40) !important;
    }

    /* Success message */
    .success-box {
        background: linear-gradient(135deg, #E8F5E9, #F1F8E9);
        border-left: 5px solid #43A047;
        border-radius: 10px;
        padding: 18px 22px;
        margin-top: 20px;
        font-weight: 600;
        color: #2E7D32;
        font-size: 1.0rem;
    }
    /* Warning message */
    .warning-box {
        background: #FFF8E1;
        border-left: 5px solid #FFD100;
        border-radius: 10px;
        padding: 16px 22px;
        margin-top: 16px;
        font-weight: 600;
        color: #7B6000;
        font-size: 0.95rem;
    }
    /* Info cards */
    .info-card {
        background: #F0F4FF;
        border-radius: 12px;
        padding: 16px 20px;
        margin-top: 18px;
        border: 1.5px solid #C5D3F5;
    }
    .info-card p {
        margin: 0;
        font-size: 0.88rem;
        color: #5DB8FD;
        font-weight: 500;
    }

    /* Radio & Checkbox Overrides */
    .stRadio label, .stCheckbox label {
        font-size: 1.1rem !important;
        color: #1A1A2E !important;
    }
    .stTextInput > label, .stTextArea > label {
        font-weight: 600 !important;
        color: #1A1A2E !important;
    }

    /* Sidebar Admin styling */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #5DB8FD 0%, #5DB8FD 100%) !important;
    }
    section[data-testid="stSidebar"] * {
        color: #fff !important;
    }
    section[data-testid="stSidebar"] .stTextInput > label {
        color: #FFD100 !important;
        font-weight: 700 !important;
    }
    section[data-testid="stSidebar"] .stButton > button {
        background: #FFD100 !important;
        color: #5DB8FD !important;
        border-radius: 8px !important;
    }

    /* Streamlit default overrides */
    div[data-baseweb="radio"] > div {
        gap: 10px;
    }
    .stCheckbox {
        margin-bottom: 4px;
    }
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ─── Constants ───────────────────────────────────────────────────────────────────
EXCEL_FILE = "responses.xlsx"
ADMIN_PASSWORD = "clorox_admin_2026"
COLUMNS = ["Timestamp", "Name", "Wants Food", "Vegan", "Vegetarian",
           "Gluten-Free", "Dairy-Free", "Other Allergies"]

# ─── Excel Helpers ───────────────────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Responses"
        ws.append(COLUMNS)
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="5DB8FD")
        header_font = Font(color="FFD100", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        wb.save(EXCEL_FILE)

def load_responses():
    init_excel()
    df = pd.read_excel(EXCEL_FILE, engine="openpyxl")
    return df

def save_response(name, wants_food, vegan, vegetarian, gluten_free, dairy_free, other):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        timestamp,
        name.strip(),
        "Yes" if wants_food else "No",
        "Yes" if vegan else "No",
        "Yes" if vegetarian else "No",
        "Yes" if gluten_free else "No",
        "Yes" if dairy_free else "No",
        other.strip() if other else "None"
    ])
    wb.save(EXCEL_FILE)

def name_already_submitted(name):
    df = load_responses()
    if df.empty or "Name" not in df.columns:
        return False
    return name.strip().lower() in df["Name"].str.lower().values

# ─── Initialize Excel ─────────────────────────────────────────────────────────
init_excel()

# ─── Sidebar: Admin Panel ────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔐 Admin Access")
    st.markdown("---")
    pwd = st.text_input("Enter Admin Password", type="default", key="admin_pwd")
    admin_btn = st.button("Log In")

    if "admin_logged_in" not in st.session_state:
        st.session_state.admin_logged_in = False

    if admin_btn:
        if pwd == ADMIN_PASSWORD:
            st.session_state.admin_logged_in = True
        else:
            st.session_state.admin_logged_in = False
            st.error("❌ Incorrect password.")

    if st.session_state.admin_logged_in:
        st.success("✅ Access granted!")
        st.markdown("---")
        st.markdown("### 📋 Responses")
        df = load_responses()
        if df.empty:
            st.info("No responses yet.")
        else:
            st.dataframe(df, use_container_width=True)
            st.markdown(f"**Total submissions:** {len(df)}")
            # Download Excel
            with open(EXCEL_FILE, "rb") as f:
                st.download_button(
                    label="⬇️ Download Excel",
                    data=f,
                    file_name="meeting_responses.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ─── Main Page ──────────────────────────────────────────────────────────────────

# Hero Banner
st.markdown("""
<div class="hero-banner">
    <div class="hero-tag">📅 Upcoming Event</div>
    <h1 class="hero-title">A&DS Monthly Meeting</h1>
    <p class="hero-subtitle">Help us plan ahead, let us know if you'd like food at the meeting!</p>
</div>
""", unsafe_allow_html=True)

# Form Card
st.markdown('<div class="form-card">', unsafe_allow_html=True)

st.markdown('<span class="section-label">👤 Your Information</span>', unsafe_allow_html=True)
name = st.text_input("Full Name", placeholder="e.g., Casey Maldonado", key="name_input")

st.markdown('<hr class="divider">', unsafe_allow_html=True)

st.markdown('<span class="section-label">🍽️ Food Preference</span>', unsafe_allow_html=True)
wants_food = st.radio(
    "Would you like food at the meeting?",
    options=["Yes, please!", "No, thank you"],
    index=None,
    key="food_pref"
)

# Dietary restrictions — only show if they want food
vegan = vegetarian = gluten_free = dairy_free = False
other_allergies = ""

if wants_food == "Yes, please!":
    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    st.markdown('<span class="section-label">🥗 Dietary Restrictions</span>', unsafe_allow_html=True)
    st.markdown("*Select all that apply:*")

    col1, col2 = st.columns(2)
    with col1:
        vegan = st.checkbox("Vegan")
        gluten_free = st.checkbox("Gluten-Free")
    with col2:
        vegetarian = st.checkbox("Vegetarian")
        dairy_free = st.checkbox("Dairy-Free")

    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    st.markdown('<span class="section-label">⚠️ Other Allergies</span>', unsafe_allow_html=True)
    other_allergies = st.text_area(
        "Any other allergies or dietary notes?",
        placeholder="e.g., nut allergy, shellfish, etc. Leave blank if none.",
        height=90,
        key="other_allergies"
    )

st.markdown('<hr class="divider">', unsafe_allow_html=True)

# Submit button
submit = st.button("Submit Response", key="submit_btn")

st.markdown('</div>', unsafe_allow_html=True)  # close form-card

# ─── Handle Submission ───────────────────────────────────────────────────────────
if submit:
    if not name.strip():
        st.markdown('<div class="warning-box">⚠️ Please enter your full name before submitting.</div>',
                    unsafe_allow_html=True)
    elif wants_food is None:
        st.markdown('<div class="warning-box">⚠️ Please select your food preference before submitting.</div>',
                    unsafe_allow_html=True)
    elif name_already_submitted(name):
        st.markdown(f'''<div class="warning-box">
            ⚠️ A response for <strong>{name.strip()}</strong> has already been submitted.
            Please contact the organizer if you need to make changes.
        </div>''', unsafe_allow_html=True)
    else:
        food_bool = wants_food == "Yes, please!"
        save_response(name, food_bool, vegan, vegetarian, gluten_free, dairy_free, other_allergies)
        st.markdown(f'''<div class="success-box">
            Thank you, <strong>{name.strip()}</strong>! Your response has been recorded.
            See you at the A&DS Monthly Meeting!
        </div>''', unsafe_allow_html=True)
